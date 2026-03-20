"""
Módulo de Extracción de Riegos.
Wrapper de la lógica de extract_riego.py para uso en Streamlit.
"""
import pandas as pd
import io
from datetime import datetime, timedelta
import openpyxl

# Constantes del módulo original
EQ_SHEETS = [
    "Eq1","Eq2","Eq3","Eq4","Eq5","Eq6","Eq7","Eq9",
    "Eq10","Eq11","Eq12","Eq13","Eq14","Eq15","Eq16","Eq17",
    "Eq18","Eq19","Eq20","Eq21","Eq22"
]

FERT_NAMES = {
    "Sulfato Zinc","Nitrato Amonio","Nitrato calcio","Nitrato Calcio",
    "Cloruro potasio","Cloruro Potasio","Acido Borico",
    "Sulfato magnesio","FMA","Urea","Novatec",
    "Nitrato Potasio","Sulfato Potasio"
}

DIAS_SEMANA = {
    0: "lun", 1: "mar", 2: "mie", 3: "jue", 4: "vie", 5: "sab", 6: "dom"
}


def parse_dates(date_str: str) -> list:
    """Parse single date or date range (YYYY-MM-DD or YYYY-MM-DD:YYYY-MM-DD)."""
    if ":" in date_str:
        start_s, end_s = date_str.split(":")
        start = datetime.strptime(start_s.strip(), "%Y-%m-%d")
        end = datetime.strptime(end_s.strip(), "%Y-%m-%d")
    else:
        start = datetime.strptime(date_str.strip(), "%Y-%m-%d")
        end = start
    dates = []
    current = start
    while current <= end:
        dates.append(current)
        current += timedelta(days=1)
    return dates


def parse_eq_number(sheet_name: str) -> int:
    """Extract equipment number from sheet name like 'Eq1' -> 1."""
    return int(sheet_name.replace("Eq", ""))


def discover_sheet_structure(ws):
    """Discover the structure of an EqX sheet."""
    sector_header_row = None
    hrs_header_row = None
    
    for r in range(5, 9):
        row_data = list(ws.iter_rows(min_row=r, max_row=r, values_only=False))[0]
        row_vals = {c.column: c.value for c in row_data if c.value is not None}
        
        has_sector = any("SECTOR" in str(v) for v in row_vals.values())
        has_hrs = any(str(v).strip() == "Hrs" for v in row_vals.values())
        
        if has_sector and sector_header_row is None:
            sector_header_row = r
        if has_hrs and hrs_header_row is None:
            hrs_header_row = r
    
    if hrs_header_row is None:
        return None
    
    hrs_row = list(ws.iter_rows(min_row=hrs_header_row, max_row=hrs_header_row, values_only=False))[0]
    hrs_cols = []
    all_cols = {}
    for c in hrs_row:
        if c.value is not None:
            all_cols[c.column] = str(c.value).strip()
            if str(c.value).strip() == "Hrs":
                hrs_cols.append(c.column)
    
    dia_row = None
    for r in range(hrs_header_row - 1, hrs_header_row + 2):
        row_data = list(ws.iter_rows(min_row=r, max_row=r, values_only=False))[0]
        for c in row_data:
            if c.value is not None and str(c.value).strip() == "Dia":
                dia_row = r
                break
        if dia_row:
            break
    
    data_start_row = (dia_row + 1) if dia_row else (hrs_header_row + 1)
    
    sectors = {}
    for i, hrs_col in enumerate(hrs_cols):
        sector_num = i + 1
        next_hrs = hrs_cols[i + 1] if i + 1 < len(hrs_cols) else hrs_col + 20
        
        fert_cols = []
        m3_col = None
        for col in range(hrs_col + 1, min(next_hrs, hrs_col + 16)):
            label = all_cols.get(col, "")
            if label in FERT_NAMES:
                fert_cols.append(col)
            elif label == "M3":
                m3_col = col
        
        sectors[sector_num] = {
            'hrs_col': hrs_col,
            'fert_cols': fert_cols,
            'm3_col': m3_col,
        }
    
    return {
        'sectors': sectors,
        'data_start_row': data_start_row,
    }


def load_maestro(wb):
    """Load the Maestro sheet into a dict keyed by (equipo, sector)."""
    ws = wb["Maestro"]
    maestro = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        clave = row[0]
        if clave is None:
            continue
        equipo = row[1]
        sector = row[2]
        has = row[3] if row[3] is not None else 0
        variedad = row[4] if row[4] is not None else ""
        prod_est = row[5] if row[5] is not None else ""
        m3_ha_hr = row[6] if row[6] is not None else 0

        maestro[(int(equipo), int(sector))] = {
            "clave": str(clave),
            "has": float(has) if has else 0,
            "variedad": str(variedad),
            "prod_est": str(prod_est),
            "m3_ha_hr": float(m3_ha_hr) if m3_ha_hr else 0,
        }
    return maestro


def get_tipo_riego(maestro_entry: dict) -> str:
    """Determine tipo de riego from maestro data."""
    var = maestro_entry["variedad"].lower()
    prod = str(maestro_entry["prod_est"]).lower()

    if any(x in var for x in ["arbequina", "arbosana", "korinenki"]):
        if "s. tree" in prod or "s.tree" in prod or "smart" in prod:
            return "Olivo S. tree"
        return "Olivo"
    if "giffoni" in var:
        return "Avellano"
    if any(x in var for x in ["santina", "lapins", "sweet"]):
        return "Cerezo"
    if "olivo" in prod:
        return "Olivo"
    if "avellano" in prod:
        return "Avellano"
    if "cerezo" in prod:
        return "Cerezo"
    return maestro_entry["variedad"] or "Otro"


def get_fundo_caseta(eq: int) -> tuple:
    """Determine fundo and caseta from equipo number."""
    fundo_map = {
        1: ("DA", "5000"), 2: ("DA", "10000"), 3: ("DA", "20000"),
        4: ("DA", "20000"), 5: ("DA", "20000"), 6: ("DA", "20000"),
        7: ("DA", "20000"), 9: ("DA", "8x8"),
        10: ("DJ", "Embalse"), 11: ("DJ", "Embalse"),
        12: ("DJ", "Embalse"), 13: ("DJ", "Embalse"),
        14: ("DJ", "Embalse"), 15: ("DJ", "Embalse"),
        16: ("DJ", "Embalse"), 17: ("DJ", "Embalse"),
        18: ("DJ", "Embalse"), 19: ("DJ", "Embalse"),
        20: ("DJ", "Embalse"), 21: ("DJ", "Embalse"),
        22: ("DJ", "Embalse"),
    }
    return fundo_map.get(eq, ("", ""))


def extract_from_eq_sheet(wb, sheet_name: str, target_dates: list) -> list:
    """Extract irrigation data from a single EqX sheet for the target dates."""
    if sheet_name not in wb.sheetnames:
        return []
    
    ws = wb[sheet_name]
    eq_num = parse_eq_number(sheet_name)
    
    structure = discover_sheet_structure(ws)
    if structure is None:
        return []
    
    target_date_set = set(d.date() for d in target_dates)
    results = []
    
    row3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=False))[0]
    
    for row in ws.iter_rows(min_row=structure['data_start_row'], values_only=False):
        fecha_cell = row[0]
        if fecha_cell.value is None:
            continue
        
        if isinstance(fecha_cell.value, datetime):
            row_date = fecha_cell.value.date()
        else:
            continue
        
        if row_date not in target_date_set:
            continue
        
        for sec_num, sec_info in structure['sectors'].items():
            hrs_col = sec_info['hrs_col']
            hrs_val = row[hrs_col - 1].value if hrs_col - 1 < len(row) else None
            
            if hrs_val is None or hrs_val == 0:
                continue
            
            try:
                hours = float(hrs_val)
            except (ValueError, TypeError):
                continue
            
            if hours <= 0:
                continue
            
            # Get M3 from planilla
            m3_planilla = None
            if sec_info['m3_col']:
                m3_idx = sec_info['m3_col'] - 1
                if m3_idx < len(row) and row[m3_idx].value is not None:
                    try:
                        m3_planilla = float(row[m3_idx].value)
                    except (ValueError, TypeError):
                        pass
            
            # Check fertilizantes
            con_fert = "No"
            for fc in sec_info['fert_cols']:
                if fc - 1 < len(row):
                    fv = row[fc - 1].value
                    if fv is not None and fv != 0:
                        try:
                            if float(fv) > 0:
                                con_fert = "Si"
                                break
                        except (ValueError, TypeError):
                            pass
            
            results.append({
                'equipo': eq_num,
                'sector': sec_num,
                'fecha': row_date,
                'horas': hours,
                'm3_planilla': m3_planilla,
                'con_fert': con_fert,
            })
    
    return results


def process_extraction_streamlit(
    uploaded_file: bytes,
    date_str: str,
    log_callback=print
) -> pd.DataFrame:
    """
    Procesa Excel uploadado y retorna DataFrame con datos extraídos.
    
    Args:
        uploaded_file: Contenido del archivo Excel (bytes)
        date_str: String de fecha/rango (ej: "2025-03-10" o "2025-03-10:2025-03-14")
        log_callback: Función para mostrar logs (default: print)
    
    Returns:
        DataFrame con columnas: equipo, sector, fecha, horas, m3_planilla, con_fert
    """
    log_callback(f"Procesando planilla para rango: {date_str}")
    
    # Cargar workbook desde bytes
    try:
        wb = openpyxl.load_workbook(io.BytesIO(uploaded_file), data_only=True)
    except Exception as e:
        log_callback(f"Error al abrir Excel: {e}")
        raise ValueError(f"Error al abrir archivo Excel: {e}")
    
    # Verificar que tenga hojas EqX
    available_sheets = [s for s in EQ_SHEETS if s in wb.sheetnames]
    if not available_sheets:
        raise ValueError("El archivo no contiene hojas de equipos (Eq1, Eq2, etc.)")
    
    log_callback(f"Hojas disponibles: {', '.join(available_sheets)}")
    
    # Parsear fechas
    target_dates = parse_dates(date_str)
    log_callback(f"Fechas a procesar: {[d.strftime('%Y-%m-%d') for d in target_dates]}")
    
    # Extraer datos
    all_data = []
    for sheet_name in available_sheets:
        entries = extract_from_eq_sheet(wb, sheet_name, target_dates)
        if entries:
            log_callback(f"  {sheet_name}: {len(entries)} registros")
        all_data.extend(entries)
    
    wb.close()
    
    if not all_data:
        log_callback("No se encontraron datos de riego para las fechas indicadas.")
        return pd.DataFrame()
    
    log_callback(f"Total: {len(all_data)} registros extraídos")
    
    # Convertir a DataFrame
    df = pd.DataFrame(all_data)
    
    # Renombrar columnas a español y agregar columnas derivadas
    def get_fundo(eq):
        fundo_map = {
            1: "DA", 2: "DA", 3: "DA", 4: "DA", 5: "DA", 6: "DA", 7: "DA", 9: "DA",
            10: "DJ", 11: "DJ", 12: "DJ", 13: "DJ", 14: "DJ", 15: "DJ", 16: "DJ", 17: "DJ",
            18: "DJ", 19: "DJ", 20: "DJ", 21: "DJ", 22: "DJ"
        }
        return fundo_map.get(eq, "")
    
    # Renombrar columnas existentes (no crear nuevas)
    df = df.rename(columns={
        'fecha': 'Fecha',
        'horas': 'Horas',
        'm3_planilla': 'M3',
        'con_fert': 'Con Fertilizante'
    })
    
    # Agregar columnas derivadas
    df['Fundo'] = df['equipo'].apply(get_fundo)
    df['Nombre Sector'] = "E" + df['equipo'].astype(str) + "S" + df['sector'].astype(str)
    
    # Convertir M3 a número (evitar NaN)
    df['M3'] = pd.to_numeric(df['M3'], errors='coerce').fillna(0)
    
    return df
