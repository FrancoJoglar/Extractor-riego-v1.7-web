"""
Módulo de Programación de Horarios.
Wrapper de la lógica de programar_horarios.py para uso en Streamlit.
"""
import streamlit as st
from supabase import create_client, Client
import pandas as pd
from datetime import datetime


def get_supabase_client() -> Client:
    """Crea cliente Supabase desde st.secrets."""
    import os
    url = st.secrets.get("SUPABASE_URL") or os.getenv("SUPABASE_URL")
    key = st.secrets.get("SUPABASE_KEY") or os.getenv("SUPABASE_KEY")
    
    if not url or not key:
        raise ValueError("SUPABASE_URL y SUPABASE_KEY deben estar configuradas")
    
    return create_client(url, key)


def formatear_hora(hora_float: float) -> str:
    """Convierte hora float (6.5) a string HH:MM."""
    horas = int(hora_float)
    minutos = int(round((hora_float - horas) * 60))
    if horas >= 24:
        horas = horas - 24
    if horas == 0 and minutos == 0:
        minutos = 1
    return f"{horas:02d}:{minutos:02d}"


def es_fertilizante(val) -> bool:
    """Convierte valor a booleano."""
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.upper() in ['TRUE', 'VERDADERO', '1', 'YES', 'S', 'SI']
    return bool(val)


def calcular_horario(equipo_df: pd.DataFrame) -> dict:
    """
    Calcula horarios para un equipo con la lógica:
    1. Sectores CON fert: 06:00 - 16:00 (SECUENCIAL)
    2. Sectores SIN fert: después de CON fert
    3. Si total > 14h: cálculo hacia atrás desde 06:00 para SIN fert
    """
    resultados = {}
    hora_termino_fert = 6.0
    
    equipo_df = equipo_df.copy()
    equipo_df['con_fertilizante'] = equipo_df['con_fertilizante'].apply(es_fertilizante)
    
    total_horas = equipo_df['horas_solicitadas'].sum()
    
    # Extraer número de sector
    if 'sector_nombre' in equipo_df.columns:
        equipo_df['sector_num'] = equipo_df['sector_nombre'].str.extract(r'S(\d+)').astype(float)
    else:
        equipo_df['sector_num'] = 1
    
    sectores_con_fert = equipo_df[equipo_df['con_fertilizante'] == True].sort_values('sector_num')
    sectores_sin_fert = equipo_df[equipo_df['con_fertilizante'] == False].sort_values('sector_num')
    
    tiene_fertilizante = len(sectores_con_fert) > 0
    
    # 1. PROGRAMAR CON FERTILIZACIÓN
    if len(sectores_con_fert) > 0:
        horas_por_sector = []
        for idx, row in sectores_con_fert.iterrows():
            horas_por_sector.append({'idx': idx, 'horas': row['horas_solicitadas'], 'sector': row.get('sector_nombre', f'S{row.get("sector_num", "?")}')})
        
        hora_actual = 6.0
        for item in horas_por_sector:
            hora_fin = hora_actual + item['horas']
            resultados[item['idx']] = {
                'Hora Inicio': formatear_hora(hora_actual),
                'Tipo': "Horario Inicio" if hora_actual == 6.0 else "Secuencial"
            }
            hora_actual += item['horas']
        
        hora_termino_fert = hora_actual
    
    # 2. PROGRAMAR SIN FERTILIZACIÓN
    if len(sectores_sin_fert) > 0:
        if tiene_fertilizante:
            hora_actual = hora_termino_fert
        elif total_horas > 14:
            primera_hora = sectores_sin_fert['horas_solicitadas'].iloc[0]
            hora_actual = 6.0 - primera_hora
            if hora_actual < 0:
                hora_actual = 0
        else:
            hora_actual = 6.0
        
        primer_sector = len(resultados) == 0
        
        for idx, row in sectores_sin_fert.iterrows():
            horas_req = row['horas_solicitadas']
            resultados[idx] = {
                'Hora Inicio': formatear_hora(hora_actual),
                'Tipo': 'Horario Inicio' if primer_sector else 'Secuencial'
            }
            hora_actual += horas_req
            primer_sector = False
    
    return resultados


def get_schedule_data(fecha: str, log_callback=print) -> pd.DataFrame:
    """
    Obtiene datos de riegos pendientes de Supabase para una fecha.
    
    Args:
        fecha: Fecha en formato YYYY-MM-DD
        log_callback: Función para logs
    
    Returns:
        DataFrame con los datos
    """
    try:
        supabase = get_supabase_client()
        log_callback(f"Obteniendo datos para fecha: {fecha}")
        
        # Convertir YYYY-MM-DD a DD-MM-YYYY para Supabase
        fecha_parts = fecha.split('-')
        fecha_supabase = f"{fecha_parts[2]}-{fecha_parts[1]}-{fecha_parts[0]}"
        
        response = supabase.table("vista_riegos_solicitados").select("*").eq("estado", "pendiente").eq("fecha_solicitado", fecha_supabase).execute()
        
        if not response.data:
            log_callback(f"No hay riegos para la fecha {fecha}")
            return pd.DataFrame()
        
        df = pd.DataFrame(response.data)
        log_callback(f"Encontrados {len(df)} registros")
        return df
        
    except Exception as e:
        log_callback(f"Error al obtener datos: {e}")
        raise


def generate_schedule(fecha: str, log_callback=print) -> pd.DataFrame:
    """
    Genera programación de horarios para una fecha.
    
    Args:
        fecha: Fecha en formato YYYY-MM-DD
        log_callback: Función para logs
    
    Returns:
        DataFrame con la programación calculada
    """
    # Obtener datos
    df = get_schedule_data(fecha, log_callback)
    
    if df.empty:
        return df
    
    df = df.copy()
    df['Hora Inicio'] = ""
    df['Tipo Programación'] = ""
    
    # Extraer números de equipo y sector
    if 'equipo_nombre' in df.columns:
        df['equipo_num'] = df['equipo_nombre'].str.extract(r'(\d+)').astype(float)
    if 'sector_nombre' in df.columns:
        df['sector_num'] = df['sector_nombre'].str.extract(r'S(\d+)').astype(float)
    
    # Procesar por equipo
    equipos = df['equipo_num'].unique()
    
    for equipo in sorted(equipos):
        equipo_df = df[df['equipo_num'] == equipo]
        resultados = calcular_horario(equipo_df)
        
        for idx, row in equipo_df.iterrows():
            if idx in resultados:
                df.loc[idx, 'Hora Inicio'] = resultados[idx]['Hora Inicio']
                df.loc[idx, 'Tipo Programación'] = resultados[idx]['Tipo']
    
    # Ordenar por hora de inicio
    df = df.sort_values(by=['equipo_num', 'sector_num'])
    
    return df


def apply_excel_styles(df: pd.DataFrame, output):
    """
    Aplica estilos corporativos al Excel exportado.
    
    Args:
        df: DataFrame con los datos
        output: Ruta del archivo Excel de salida (str) o BytesIO buffer
    """
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io
    
    is_buffer = isinstance(output, io.BytesIO)
    
    # Guardar DataFrame a Excel
    if is_buffer:
        # Guardar a bytes primero, luego cargar
        df.to_excel(output, index=False, engine='openpyxl')
        output.seek(0)
        wb = openpyxl.load_workbook(output)
    else:
        df.to_excel(output, index=False)
        wb = openpyxl.load_workbook(output)
    ws = wb.active
    
    # ============================================
    # 1. CONFIGURACIÓN DE COLORES Y ESTILOS
    # ============================================
    
    # Encabezado: Azul #0066CC (formato ARGB)
    AZUL_ENCABEZADO = "FF0066CC"
    header_fill = PatternFill(start_color=AZUL_ENCABEZADO, end_color=AZUL_ENCABEZADO, fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    
    # Paleta 2 tonos azul por equipo (alternados, formato ARGB con alpha FF)
    AZUL_1 = "FF5B9BD5"  # Azul corporativo
    AZUL_2 = "FFBDD7EE"  # Azul claro
    COLORES_EQUIPOS = [AZUL_1, AZUL_2]
    
    # Mapear equipos a colores (alternar)
    # La columna puede ser 'equipo', 'Equipo', o 'equipo_nombre'
    equipos_col = 'equipo' if 'equipo' in df.columns else ('Equipo' if 'Equipo' in df.columns else ('equipo_nombre' if 'equipo_nombre' in df.columns else None))
    equipos_unicos = df[equipos_col].unique() if equipos_col else []
    color_por_equipo = {eq: COLORES_EQUIPOS[i % 2] for i, eq in enumerate(sorted(equipos_unicos))}
    
    # Fuentes
    font_dato = Font(name="Calibri", size=11, bold=False)
    font_importante = Font(name="Calibri", size=14, bold=True)  # Sector, M3
    font_hora = Font(name="Calibri", size=12, bold=True, color="FFFF6600")  # Hora Inicio NARANJO (ARGB)
    
    # Bordes suaves
    border_side = Side(color="FFD3D3D3", style='thin')
    border = Border(left=border_side, right=border_side, top=border_side, bottom=border_side)
    
    # ============================================
    # 2. MAPEO DE COLUMNAS
    # ============================================
    columnas = {cell.value: idx for idx, cell in enumerate(ws[1])}
    
    # Columnas importantes - buscar múltiples variantes de nombres
    col_sector = columnas.get('Nombre Sector', columnas.get('Sector', columnas.get('sector', None)))
    col_m3 = columnas.get('M3', columnas.get('m3', None))
    col_hora = columnas.get('Hora Inicio', columnas.get('hora_inicio', None))
    col_equipo = columnas.get('Equipo', columnas.get('equipo', columnas.get('equipo_nombre', None)))
    
    # NOTA: Las columnas ya vienen filtradas desde la página (Fecha, Equipo, Sector, etc.)
    # No se ocultan columnas aquí
    
    # ============================================
    # 3. APLICAR ESTILOS AL ENCABEZADO
    # ============================================
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Redondear M3 a entero (estándar: >0.5 hacia arriba, <0.5 hacia abajo)
    if col_m3 is not None:
        for row in ws.iter_rows(min_row=2):
            cell = row[col_m3]
            if cell.value is not None:
                try:
                    cell.value = int(cell.value + 0.5) if cell.value >= 0 else int(cell.value - 0.5)
                except (ValueError, TypeError):
                    pass
    
    # ============================================
    # 4. APLICAR ESTILOS A LAS FILAS
    # ============================================
    for row in ws.iter_rows(min_row=2):
        equipo = row[col_equipo].value if col_equipo is not None else None
        fill_color = color_por_equipo.get(equipo, "FFFFFF")
        fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
        
        for col_idx, cell in enumerate(row):
            cell.fill = fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = border
            
            # Hora Inicio en NARANJA + negrita
            if col_idx == col_hora:
                cell.font = font_hora
            # Sector y M3 en negrita
            elif col_idx in [col_sector, col_m3]:
                cell.font = font_importante
            else:
                cell.font = font_dato
    
    # ============================================
    # 5. AJUSTAR ANCHOS DE COLUMNAS
    # ============================================
    anchos = {
        'A': 16,  # Fecha
        'B': 12,  # Equipo
        'C': 10,  # Sector
        'D': 15,  # Jefe de Campo
        'E': 8,   # Horas
        'F': 14,  # M3
        'G': 18,  # Con Fertilizante
        'H': 14,  # Hora Inicio
        'I': 18,  # Tipo Programación
    }
    for col_letter, width in anchos.items():
        if col_letter in ws.column_dimensions:
            ws.column_dimensions[col_letter].width = width
    
    wb.save(output)
