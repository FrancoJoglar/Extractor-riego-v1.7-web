import pandas as pd
from supabase import create_client, Client
import os
from datetime import datetime, timedelta

SUPABASE_URL = "https://miikjrfqmmkzknyngwen.supabase.co"
SUPABASE_KEY = "sb_publishable_tixCrY6poQrVDEzeRuDhow_ND7tvVCQ"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

def obtener_datos(fecha_filtro=None):
    print("Obteniendo datos de riegos pendientes...")
    
    query = supabase.table("vista_riegos_solicitados").select("*").eq("estado", "pendiente")
    
    if fecha_filtro:
        print(f"Filtrando por fecha: {fecha_filtro}")
        query = query.eq("fecha_solicitado", fecha_filtro)
    
    response = query.execute()
    data = response.data
    
    if not data:
        print(f"No hay riegos pendientes para programar{' para la fecha ' + fecha_filtro if fecha_filtro else ''}.")
        return None
    
    df = pd.DataFrame(data)
    print(f"Se encontraron {len(df)} riegos pendientes")
    return df


def extraer_numeros(df):
    if 'equipo_nombre' in df.columns:
        df['equipo_num'] = df['equipo_nombre'].str.extract(r'(\d+)').astype(float)
    if 'sector_nombre' in df.columns:
        df['sector_num'] = df['sector_nombre'].str.extract(r'S(\d+)').astype(float)
    return df


def calcular_hora_base(total_horas_equipo):
    return 4 if total_horas_equipo > 16 else 6


def ajustar_para_fertilizante(hora_actual, horas_requeridas, tiene_fertilizante):
    if not tiene_fertilizante:
        return hora_actual
    
    hora_termino = hora_actual + horas_requeridas
    
    if hora_termino <= 16:
        return hora_actual
    
    nueva_hora_inicio = 16 - horas_requeridas
    if nueva_hora_inicio < 4:
        nueva_hora_inicio = 4
    
    return nueva_hora_inicio


def formatear_hora(hora_float):
    horas = int(hora_float)
    minutos = int((hora_float - horas) * 60)
    return f"{horas:02d}:{minutos:02d}"


def programar_horarios_por_fecha_equipo(df):
    df = df.copy()
    df['Hora Inicio'] = ""
    df['Tipo Programación'] = ""
    
    df = extraer_numeros(df)
    
    if 'fecha_solicitado' not in df.columns or 'equipo_num' not in df.columns:
        print("ERROR: Faltan columnas necesarias")
        return df
    
    fechas = df['fecha_solicitado'].unique()
    
    for fecha in sorted(fechas):
        fecha_df = df[df['fecha_solicitado'] == fecha]
        equipos = fecha_df['equipo_num'].unique()
        
        for equipo in sorted(equipos):
            equipo_df = fecha_df[fecha_df['equipo_num'] == equipo].sort_values('sector_num')
            
            total_horas = equipo_df['horas_solicitadas'].sum()
            hora_base = calcular_hora_base(total_horas)
            
            hora_actual = hora_base
            
            tiene_fert = equipo_df['con_fertilizante'].any()
            if tiene_fert:
                horas_fert = equipo_df[equipo_df['con_fertilizante'] == True]['horas_solicitadas'].sum()
                hora_ajustada = 16 - horas_fert
                if hora_ajustada < 4:
                    hora_ajustada = 4
                hora_actual = hora_ajustada
            
            print(f"Fecha {fecha.strftime('%Y-%m-%d') if isinstance(fecha, (datetime, pd.Timestamp)) else fecha} | Equipo {int(equipo)}: hora_base={hora_base}h (total: {total_horas}h)")
            
            primer_sector = True
            sector_anterior = None
            
            for idx, row in equipo_df.iterrows():
                tiene_fert = row.get('con_fertilizante', False)
                
                hora_actual = ajustar_para_fertilizante(hora_actual, row['horas_solicitadas'], tiene_fert)
                
                hora_str = formatear_hora(hora_actual)
                df.loc[idx, 'Hora Inicio'] = hora_str
                
                if primer_sector:
                    df.loc[idx, 'Tipo Programación'] = "Horario Inicio"
                    primer_sector = False
                else:
                    df.loc[idx, 'Tipo Programación'] = "Secuencial"
                
                sector_anterior = row['sector_nombre']
                
                print(f"  {row['sector_nombre']}: {hora_str} ({df.loc[idx, 'Tipo Programación']}) (fert: {tiene_fert}, {row['horas_solicitadas']}h)")
                
                hora_actual += row['horas_solicitadas']
    
    return df


def guardar_excel(df, output_path=None, nombre_base="Planilla_Programacion"):
    cols_a_borrar = ['equipo_num', 'sector_num']
    df = df.drop(columns=[c for c in cols_a_borrar if c in df.columns], errors='ignore')
    
    if 'fecha_solicitado' in df.columns:
        if isinstance(df['fecha_solicitado'].iloc[0], (datetime, pd.Timestamp)):
            df['fecha_solicitado'] = df['fecha_solicitado'].dt.strftime('%d-%m-%Y')
    
    if 'm3_estimados' in df.columns:
        df['m3_estimados'] = df['m3_estimados'].fillna(0).round(0).astype(int)
    
    columnas_orden = [
        'fecha_solicitado', 
        'equipo_nombre', 
        'sector_nombre', 
        'jefe_campo', 
        'horas_solicitadas', 
        'm3_estimados', 
        'con_fertilizante',
        'Hora Inicio',
        'Tipo Programación'
    ]
    cols_final = [c for c in columnas_orden if c in df.columns]
    df = df[cols_final]
    
    nombres = {
        'fecha_solicitado': 'Fecha Solicitado',
        'equipo_nombre': 'Equipo',
        'sector_nombre': 'Sector',
        'jefe_campo': 'Jefe de Campo',
        'horas_solicitadas': 'Horas',
        'm3_estimados': 'M3 Estimados',
        'con_fertilizante': 'Con Fertilizante',
        'Hora Inicio': 'Hora Inicio',
        'Tipo Programación': 'Tipo Programación'
    }
    df = df.rename(columns=nombres)
    
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"{nombre_base}_{timestamp}.xlsx"
        output_path = os.path.join(os.getcwd(), output_filename)
    
    df.to_excel(output_path, index=False)
    
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        azul_encabezado = "0070C0"
        header_fill = PatternFill(start_color=azul_encabezado, end_color=azul_encabezado, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF")
        
        naranjo = "FF6600"
        hora_inicio_font = Font(color=naranjo, bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        colores_fondo = ["F2F7FB", "E1EEF8"]
        
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
            row_fill = PatternFill(start_color=colores_fondo[i % 2], end_color=colores_fondo[i % 2], fill_type="solid")
            for idx, cell in enumerate(row, 1):
                cell.fill = row_fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                if idx in [3, 6]:
                    cell.font = Font(bold=True)
                
                if idx == 8:
                    cell.font = hora_inicio_font
        
        wb.save(output_path)
        print("Estilos aplicados correctamente.")
    except Exception as e:
        print(f"Error al aplicar estilos: {e}")
    
    print(f"\nArchivo generado: {output_path}")
    os.startfile(output_path)


def programar_horarios_inicio(fecha_filtro=None, output_path=None):
    df = obtener_datos(fecha_filtro)
    
    if df is None or df.empty:
        return
    
    if 'fecha_solicitado' in df.columns:
        df['fecha_solicitado'] = pd.to_datetime(df['fecha_solicitado'], dayfirst=True, errors='coerce')
    
    if 'equipo_nombre' in df.columns:
        df['equipo_num'] = df['equipo_nombre'].str.extract(r'(\d+)').astype(float)
    if 'sector_nombre' in df.columns:
        df['sector_num'] = df['sector_nombre'].str.extract(r'S(\d+)').astype(float)
    
    df = df.sort_values(by=['fecha_solicitado', 'equipo_num', 'sector_num'])
    
    df_programado = programar_horarios_por_fecha_equipo(df)
    
    guardar_excel(df_programado, output_path)


def programar_horarios_a_excel(fecha_filtro, output_path):
    """Función para ser llamada desde la GUI"""
    programar_horarios_inicio(fecha_filtro, output_path)


if __name__ == "__main__":
    import sys
    fecha = sys.argv[1] if len(sys.argv) > 1 else None
    programar_horarios_inicio(fecha)