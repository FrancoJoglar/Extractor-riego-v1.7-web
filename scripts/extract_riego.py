#!/usr/bin/env python3
"""
riego-extractor: Extrae datos de riego desde la planilla M3 reales.
Lee cada hoja EqX directamente + hoja Maestro, genera un Excel filtrable.

Uso:
  python extract_riego.py <ruta_planilla> <fecha_o_rango> <salida.xlsx>

Ejemplos:
  python extract_riego.py "planilla.xlsx" "2025-12-10" "salida.xlsx"
  python extract_riego.py "planilla.xlsx" "2025-12-10:2025-12-14" "salida.xlsx"
"""

import sys
import os
from datetime import datetime, timedelta, time as dtime
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd

# All equipment sheet names
EQ_SHEETS = [
    "Eq1","Eq2","Eq3","Eq4","Eq5","Eq6","Eq7","Eq9",
    "Eq10","Eq11","Eq12","Eq13","Eq14","Eq15","Eq16","Eq17",
    "Eq18","Eq19","Eq20","Eq21","Eq22"
]

FERT_NAMES = {
    "Sulfato Zinc","Nitrato Amonio","Nitrato calcio","Nitrato Calcio",
    "Cloruro potasio","Cloruro Potasio","Acido Borico",
    "Sulfato magnesio","FMA","Urea","Novatec",
    "Nitrato Potasio","Sulfato Potasio"
}

DIAS_SEMANA = {
    0: "lun", 1: "mar", 2: "mie", 3: "jue", 4: "vie", 5: "sab", 6: "dom"
}

PLANILLA_JEFES = "Planilla tipo Programacion riego.xlsx"

# --- SUPABASE CONFIG ---
SUPABASE_URL = "https://miikjrfqmmkzknyngwen.supabase.co"
SUPABASE_KEY = "sb_publishable_tixCrY6poQrVDEzeRuDhow_ND7tvVCQ"

def upload_to_supabase(df_results, log_callback=print):
    """Sube los resultados extraídos a Supabase."""
    try:
        from supabase import create_client
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        
        log_callback("Conectando con Supabase para sincronización...")
        
        # 1. Obtener Maestros para mapeo
        fundos_resp = supabase.table("fundos").select("id, nombre").execute()
        sectores_resp = supabase.table("sectores").select("id, nombre, equipo_id").execute()
        
        fundo_map = {f['nombre']: f['id'] for f in fundos_resp.data}
        sector_map = {s['nombre']: s['id'] for s in sectores_resp.data}
        sector_to_eq = {s['nombre']: s['equipo_id'] for s in sectores_resp.data}

        # 2. Preparar registros
        registros = []
        for _, row in df_results.iterrows():
            fundo_id = fundo_map.get(row['Fundo'])
            nom_sector = row['Nombre Sector']
            sector_id = sector_map.get(nom_sector)
            equipo_id = sector_to_eq.get(nom_sector)

            if fundo_id and sector_id:
                registros.append({
                    "fundo_id": fundo_id,
                    "equipo_id": equipo_id,
                    "sector_id": sector_id,
                    "fecha_solicitado": str(row['Fecha']),
                    "horas_solicitadas": float(row['Horas']),
                    "m3_estimados": float(row['M3']) if pd.notna(row['M3']) else 0,
                    "con_fertilizante": True if row['Con Fertilizante'] == "Si" else False,
                    "estado": "pendiente"
                })

        if registros:
            log_callback(f"Sincronizando {len(registros)} registros con Supabase...")
            supabase.table("riegos_solicitados").upsert(
                registros, 
                on_conflict="fecha_solicitado, sector_id, horas_solicitadas"
            ).execute()
            log_callback("¡Sincronización con Supabase exitosa!")
        else:
            log_callback("No hay registros válidos para sincronizar.")
            
    except Exception as e:
        log_callback(f"Advertencia: No se pudo sincronizar con Supabase: {e}")


def clear_riegos_solicitados(log_callback=print):
    """Borra todos los registros de la tabla riegos_solicitados."""
    try:
        from supabase import create_client
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        
        log_callback("⚠️ Conectando a Supabase para limpieza de tabla...")
        
        # Eliminar todos los registros
        # En Supabase/PostgREST para borrar todos se puede usar delete().neq('id', -1) o similar
        # si no se permite borrar sin filtro.
        # Probamos con .neq('id', -1) que suele funcionar para 'borrar todo' si id es serial.
        resp = supabase.table("riegos_solicitados").delete().neq("id", -1).execute()
        
        if hasattr(resp, 'data'):
            num_borrados = len(resp.data)
            log_callback(f"✅ Se han borrado {num_borrados} registros de 'riegos_solicitados'.")
        else:
            log_callback("✅ Tabla 'riegos_solicitados' limpiada con éxito.")
            
    except Exception as e:
        log_callback(f"❌ ERROR al limpiar tabla: {e}")
        raise e

def get_jefes_from_supabase(log_callback=print):
    """Obtiene el mapeo de Sector -> Jefe de Campo desde Supabase."""

    try:
        from supabase import create_client
        supabase = create_client(SUPABASE_URL, SUPABASE_KEY)
        
        # Query joining sectores and personas
        # PostgREST syntax for nested select: select('nombre, personas(nombre)')
        # Note: 'personas' is the relation name (FK in 'sectores' table)
        resp = supabase.table("sectores").select("nombre, personas(nombre)").execute()
        
        jefes_map = {}
        for item in resp.data:
            sector_name = item.get('nombre')
            persona = item.get('personas') # This will be a dict or None
            if sector_name and persona:
                jefe_nombre = persona.get('nombre')
                if jefe_nombre:
                    jefes_map[sector_name] = jefe_nombre
        
        if jefes_map:
            log_callback(f"  Cargados {len(jefes_map)} jefes desde Supabase.")
        return jefes_map
    except Exception as e:
        log_callback(f"  Advertencia: No se pudo cargar jefes desde Supabase: {e}")
        return {}


def parse_dates(date_str):
    """Parse single date or date range (YYYY-MM-DD or YYYY-MM-DD:YYYY-MM-DD)."""
    if ":" in date_str:
        start_s, end_s = date_str.split(":")
        start = datetime.strptime(start_s.strip(), "%Y-%m-%d")
        end = datetime.strptime(end_s.strip(), "%Y-%m-%d")
    else:
        start = datetime.strptime(date_str.strip(), "%Y-%m-%d")
        end = start
    dates = []
    current = start
    while current <= end:
        dates.append(current)
        current += timedelta(days=1)
    return dates


def parse_eq_number(sheet_name):
    """Extract equipment number from sheet name like 'Eq1' -> 1."""
    return int(sheet_name.replace("Eq", ""))


def discover_sheet_structure(ws):
    """
    Discover the structure of an EqX sheet.
    Returns: {
        'sectors': {sector_num: {'hrs_col': int, 'fert_cols': [int], 'm3_col': int}},
        'data_start_row': int,
        'has_row': int (row 3 always),
    }
    """
    # Strategy: scan rows 5-8 to find the row with "SECTOR X" labels
    # and the row with "Hrs" column headers
    sector_header_row = None
    hrs_header_row = None
    
    for r in range(5, 9):
        row_data = list(ws.iter_rows(min_row=r, max_row=r, values_only=False))[0]
        row_vals = {c.column: c.value for c in row_data if c.value is not None}
        
        # Check for SECTOR labels
        has_sector = any("SECTOR" in str(v) for v in row_vals.values())
        has_hrs = any(str(v).strip() == "Hrs" for v in row_vals.values())
        
        if has_sector and sector_header_row is None:
            sector_header_row = r
        if has_hrs and hrs_header_row is None:
            hrs_header_row = r
    
    if hrs_header_row is None:
        return None
    
    # Parse the Hrs header row to find column positions
    hrs_row = list(ws.iter_rows(min_row=hrs_header_row, max_row=hrs_header_row, values_only=False))[0]
    hrs_cols = []
    all_cols = {}
    for c in hrs_row:
        if c.value is not None:
            all_cols[c.column] = str(c.value).strip()
            if str(c.value).strip() == "Hrs":
                hrs_cols.append(c.column)
    
    # Data starts 1 row after the header row with "Dia"
    # Find the row with "Dia" label
    dia_row = None
    for r in range(hrs_header_row - 1, hrs_header_row + 2):
        row_data = list(ws.iter_rows(min_row=r, max_row=r, values_only=False))[0]
        for c in row_data:
            if c.value is not None and str(c.value).strip() == "Dia":
                dia_row = r
                break
        if dia_row:
            break
    
    data_start_row = (dia_row + 1) if dia_row else (hrs_header_row + 1)
    
    # For each sector (each Hrs column), find the fert columns and M3 column
    sectors = {}
    for i, hrs_col in enumerate(hrs_cols):
        sector_num = i + 1
        
        # Next Hrs col (or end)
        next_hrs = hrs_cols[i + 1] if i + 1 < len(hrs_cols) else hrs_col + 20
        
        # Find fert columns between hrs_col and next_hrs
        fert_cols = []
        m3_col = None
        for col in range(hrs_col + 1, min(next_hrs, hrs_col + 16)):
            label = all_cols.get(col, "")
            if label in FERT_NAMES:
                fert_cols.append(col)
            elif label == "M3":
                m3_col = col
        
        sectors[sector_num] = {
            'hrs_col': hrs_col,
            'fert_cols': fert_cols,
            'm3_col': m3_col,
        }
    
    return {
        'sectors': sectors,
        'data_start_row': data_start_row,
    }


def load_maestro(wb):
    """Load the Maestro sheet into a dict keyed by (equipo, sector)."""
    ws = wb["Maestro"]
    maestro = {}
    for row in ws.iter_rows(min_row=4, values_only=True):
        clave = row[0]
        if clave is None:
            continue
        equipo = row[1]
        sector = row[2]
        has = row[3] if row[3] is not None else 0
        variedad = row[4] if row[4] is not None else ""
        prod_est = row[5] if row[5] is not None else ""
        m3_ha_hr = row[6] if row[6] is not None else 0

        maestro[(int(equipo), int(sector))] = {
            "clave": str(clave),
            "has": float(has) if has else 0,
            "variedad": str(variedad),
            "prod_est": str(prod_est),
            "m3_ha_hr": float(m3_ha_hr) if m3_ha_hr else 0,
        }
    return maestro


import sys
if getattr(sys, 'frozen', False):
    # If frozen, look for database in the internal _MEIPASS or adjacent
    # But for our import we just need standard import since it is in same folder source
    pass
try:
    import database
except ImportError:
    # If running from script in root vs scripts folder, might need path adjustment
    sys.path.append(os.path.dirname(os.path.abspath(__file__)))
    import database

 # --- CONFIG ---
PLANILLA_JEFES = "Planilla tipo Programacion riego.xlsx"


# debug logger
LOG_FILE = "debug_log.txt"

def debug_log(msg):
    """Write debug message to file."""
    try:
        with open(LOG_FILE, "a", encoding="utf-8") as f:
            ts = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            f.write(f"[{ts}] {msg}\n")
    except:
        pass

def load_jefes_campo(wb, base_path=None, manual_path=None):
    """
    Load jefe de campo mapping.
    Priority 0: Supabase (Live data)
    Priority 1: Manual Path (if provided by GUI)
    Priority 2: 'Datos' sheet in 'wb' (Input file)
    Priority 3: Auto-discovery of 'Planilla tipo...'
    """
    debug_log("Iniciando load_jefes_campo...")
    jefes_from_excel = {}
    ws = None
    
    # 0. Supabase Priority
    sb_jefes = get_jefes_from_supabase(log_callback=debug_log)
    if sb_jefes:
        print(f"  Usando {len(sb_jefes)} jefes desde Supabase.")
        return sb_jefes

    # 1. Manual Path
    if manual_path and os.path.exists(manual_path):
        debug_log(f"  Usando ruta manual: {manual_path}")
        try:
            wb_ext = openpyxl.load_workbook(manual_path, read_only=True, data_only=True)
            if "Datos" in wb_ext.sheetnames:
                ws = wb_ext["Datos"]
                print(f"  Leyendo jefes desde MANUAL: {manual_path}")
            else:
                debug_log(f"  ERROR: Ruta manual {manual_path} no tiene hoja 'Datos'")
        except Exception as e:
             debug_log(f"  ERROR leyendo ruta manual: {e}")

    # 1. Try finding 'Datos' in the current workbook (if not found in manual)
    if not ws and "Datos" in wb.sheetnames:
        ws = wb["Datos"]
        debug_log("  Encontrada hoja 'Datos' en workbook actual.")
        print("  Leyendo jefes desde hoja 'Datos' en planilla actual...")
    
    # 2. Fallback: looking for external file (if still not found)
    if not ws:
        candidates = []
        
        # Helper to add parents
        def add_parents(path, levels=2):
            curr = path
            for _ in range(levels):
                curr = os.path.dirname(curr)
                candidates.append(os.path.join(curr, PLANILLA_JEFES))

        # Candidate 1: Based on input file path (and parents)
        if base_path:
            p = base_path if os.path.isdir(base_path) else os.path.dirname(base_path)
            candidates.append(os.path.join(p, PLANILLA_JEFES))
            add_parents(p)
        
        # Candidate 2: Current Working Directory (and parents)
        cwd = os.getcwd()
        candidates.append(os.path.join(cwd, PLANILLA_JEFES))
        add_parents(cwd)
        
        # Candidate 3: Bundle Directory (PyInstaller _MEIPASS)
        if hasattr(sys, '_MEIPASS'):
            bundle_dir = sys._MEIPASS
            candidates.append(os.path.join(bundle_dir, PLANILLA_JEFES))
            debug_log(f"  Ejecutando como BUNDLE. Temp dir: {bundle_dir}")

        # Candidate 4: Executable/Script directory (and parents)
        if getattr(sys, 'frozen', False):
            exe_dir = os.path.dirname(sys.executable)
            candidates.append(os.path.join(exe_dir, PLANILLA_JEFES))
            add_parents(exe_dir)
            debug_log(f"  Ejecutando como FROZEN. Exe dir: {exe_dir}")
        else:
            script_dir = os.path.dirname(os.path.abspath(__file__))
            candidates.append(os.path.join(script_dir, PLANILLA_JEFES))
            add_parents(script_dir)
            debug_log(f"  Ejecutando como SCRIPT. Script dir: {script_dir}")

        # Remove duplicates while preserving order
        unique_candidates = []
        for c in candidates:
            if c not in unique_candidates:
                unique_candidates.append(c)

        debug_log(f"  Candidatos de busqueda: {unique_candidates}")

        file_path = None
        for cand in unique_candidates:
            if os.path.exists(cand):
                file_path = cand
                break
        
        if file_path:
             debug_log(f"  Archivo encontrado en: {file_path}")
             try:
                wb_ext = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
                if "Datos" in wb_ext.sheetnames:
                    ws = wb_ext["Datos"]
                    print(f"  Leyenda jefes desde hoja 'Datos' en: {file_path}")
                    debug_log("  Hoja 'Datos' abierta exitosamente.")
                else:
                    debug_log(f"  Hoja 'Datos' NO encontrada en {file_path}. Hojas: {wb_ext.sheetnames}")
             except Exception as e:
                print(f"Error leyendo {file_path}: {e}")
                debug_log(f"  EXCEPCION leyendo archivo: {e}")
        else:
             print(f"  No se encontro {PLANILLA_JEFES}. Buscado en: {unique_candidates[:3]}...")
             debug_log("  NO SE ENCONTRO NINGUN ARCHIVO CANDIDATO.")

    if ws:
        try:
             # Headers: Nombre sector, Equipo, Sector, Caudal Nominal m3/h, Has, Variedad, Jefe de campo, Fertilizante
            header_map = {}
            data_start_row = 1
            
            def normalize(h): return str(h).strip().lower() if h else ""
            
            # Scan first 5 rows
            for r_idx, row in enumerate(ws.iter_rows(min_row=1, max_row=5, values_only=True), 1):
                row_vals = [normalize(c) for c in row]
                
                for i, val in enumerate(row_vals):
                    if val in ["equipo", "eq", "eq."]: header_map["equipo"] = i
                    elif val in ["sector", "sect", "sec", "sec.", "s."]: header_map["sector"] = i
                    elif "nombre" in val and "sector" in val: header_map["nombre_sector"] = i # Looking for 'Nombre Sector'
                    elif any(x in val for x in ["jefe", "encargado", "responsable", "jefe de campo"]): header_map["jefe"] = i
                
                # We prioritize 'Nombre Sector' now if requested
                if "nombre_sector" in header_map or ("equipo" in header_map and "sector" in header_map):
                    data_start_row = r_idx + 1
                    break

            ns_idx = header_map.get("nombre_sector")
            eq_idx = header_map.get("equipo")
            sec_idx = header_map.get("sector")
            jefe_idx = header_map.get("jefe")
            
            if (ns_idx is None and (eq_idx is None or sec_idx is None)):
                 print("Advertencia: No se encontraron columnas clave (Nombre Sector o Equipo+Sector)")
                 return {}

            for row in ws.iter_rows(min_row=data_start_row, values_only=True):
                if not row: continue
                try:
                    val_jefe = None
                    if jefe_idx is not None and jefe_idx < len(row):
                         val_jefe = row[jefe_idx]

                    if not val_jefe: continue
                    val_jefe = str(val_jefe).strip()
                    
                    # Strategy 1: Use 'Nombre Sector' (E1S1) directly
                    if ns_idx is not None and ns_idx < len(row):
                        key_ns = row[ns_idx]
                        if key_ns:
                            # Clean it up: " E1S1 " -> "E1S1"
                            key_clean = str(key_ns).strip().upper()
                            jefes_from_excel[key_clean] = val_jefe
                            continue # Successfully keyed by name

                    # Strategy 2: Fallback to Eq+Sector construction
                    if eq_idx is not None and sec_idx is not None:
                         eq = row[eq_idx]
                         sec = row[sec_idx]
                         # Cleanup and construct "E{eq}S{sec}" to match Strategy 1 format
                         # This acts as a normalizer
                         if isinstance(eq, str): eq = ''.join(filter(str.isdigit, eq))
                         if isinstance(sec, str): sec = ''.join(filter(str.isdigit, sec))
                         
                         if eq and sec:
                             key_constructed = f"E{int(eq)}S{int(sec)}"
                             jefes_from_excel[key_constructed] = val_jefe

                except Exception:
                    continue
            
            if jefes_from_excel:
                print(f"  Encontrados {len(jefes_from_excel)} jefes en Excel.")
                # We skip DB cache for now as we are relying on direct Key match
                return jefes_from_excel
        except Exception as e:
            print(f"Error procesando hoja 'Datos': {e}")
    
    return {}
    
    # Old DB Fallback removed from here to simplify logic flow as requested
    return {}


def extract_from_eq_sheet(wb, sheet_name, target_dates):
    """
    Extract irrigation data from a single EqX sheet for the target dates.
    Returns list of dicts with keys: equipo, sector, fecha, horas, m3_planilla, con_fert
    """
    if sheet_name not in wb.sheetnames:
        return []
    
    ws = wb[sheet_name]
    eq_num = parse_eq_number(sheet_name)
    
    structure = discover_sheet_structure(ws)
    if structure is None:
        return []
    
    target_date_set = set(d.date() for d in target_dates)
    results = []
    
    # Read has from row 3
    row3 = list(ws.iter_rows(min_row=3, max_row=3, values_only=False))[0]
    has_by_col = {c.column: c.value for c in row3 if c.value is not None and isinstance(c.value, (int, float))}
    
    for row in ws.iter_rows(min_row=structure['data_start_row'], values_only=False):
        fecha_cell = row[0]
        if fecha_cell.value is None:
            continue
        
        # Parse date
        if isinstance(fecha_cell.value, datetime):
            row_date = fecha_cell.value.date()
        else:
            continue
        
        if row_date not in target_date_set:
            continue
        
        # Check each sector
        for sec_num, sec_info in structure['sectors'].items():
            hrs_col = sec_info['hrs_col']
            
            # Get hours value (column index is 0-based in the row tuple)
            hrs_val = row[hrs_col - 1].value if hrs_col - 1 < len(row) else None
            
            if hrs_val is None or hrs_val == 0:
                continue
            
            try:
                hours = float(hrs_val)
            except (ValueError, TypeError):
                continue
            
            if hours <= 0:
                continue
            
            # Get M3 from planilla (already calculated)
            m3_planilla = None
            if sec_info['m3_col']:
                m3_idx = sec_info['m3_col'] - 1
                if m3_idx < len(row) and row[m3_idx].value is not None:
                    try:
                        m3_planilla = float(row[m3_idx].value)
                    except (ValueError, TypeError):
                        pass
            
            # Check fertilizantes
            con_fert = "No"
            for fc in sec_info['fert_cols']:
                if fc - 1 < len(row):
                    fv = row[fc - 1].value
                    if fv is not None and fv != 0:
                        try:
                            if float(fv) > 0:
                                con_fert = "Si"
                                break
                        except (ValueError, TypeError):
                            pass
            
            results.append({
                'equipo': eq_num,
                'sector': sec_num,
                'fecha': row_date,
                'horas': hours,
                'm3_planilla': m3_planilla,
                'con_fert': con_fert,
            })
    
    return results


def get_tipo_riego(maestro_entry):
    """Determine tipo de riego from maestro data."""
    var = maestro_entry["variedad"].lower()
    prod = str(maestro_entry["prod_est"]).lower()

    if any(x in var for x in ["arbequina", "arbosana", "korinenki"]):
        if "s. tree" in prod or "s.tree" in prod or "smart" in prod:
            return "Olivo S. tree"
        return "Olivo"
    if "giffoni" in var:
        return "Avellano"
    if any(x in var for x in ["santina", "lapins", "sweet"]):
        return "Cerezo"
    if "olivo" in prod:
        return "Olivo"
    if "avellano" in prod:
        return "Avellano"
    if "cerezo" in prod:
        return "Cerezo"
    return maestro_entry["variedad"] or "Otro"


def get_fundo_caseta(eq):
    """Determine fundo and caseta from equipo number."""
    fundo_map = {
        1: ("DA", "5000"), 2: ("DA", "10000"), 3: ("DA", "20000"),
        4: ("DA", "20000"), 5: ("DA", "20000"), 6: ("DA", "20000"),
        7: ("DA", "20000"), 9: ("DA", "8x8"),
        10: ("DJ", "Embalse"), 11: ("DJ", "Embalse"),
        12: ("DJ", "Embalse"), 13: ("DJ", "Embalse"),
        14: ("DJ", "Embalse"), 15: ("DJ", "Embalse"),
        16: ("DJ", "Embalse"), 17: ("DJ", "Embalse"),
        18: ("DJ", "Embalse"), 19: ("DJ", "Embalse"),
        20: ("DJ", "Embalse"), 21: ("DJ", "Embalse"),
        22: ("DJ", "Embalse"),
    }
    return fundo_map.get(eq, ("", ""))



def vincular_jefes_campo(df_main, jefes_map):
    """
    Realiza un VLOOKUP inteligente para asignar Jefes de Campo.
    df_main: DataFrame con los datos de riego.
    jefes_map: Diccionario {(equipo, sector): jefe} cargado desde Excel.
    """
    # Convert map to DataFrame for merging
    # jefes_map keys are (1, 1), etc.
    if not jefes_map:
        df_main['jefe_campo'] = "POR ASIGNAR"
        return df_main

    jefes_data = []
    
    # Detect key type in jefes_map
    first_key = next(iter(jefes_map))
    is_string_key = isinstance(first_key, str)

    if is_string_key:
        debug_log("  Jefes Map usa claves STRING (E1S1). Adaptando...")
        for key_sector, jefe in jefes_map.items():
            # key_sector is "E1S1"
            jefes_data.append({'key_sector': key_sector, 'jefe_lookup': jefe})
            # We don't need 'equipo'/'sector' columns for df_jefes strictly if we merge on key_sector
    else:
        debug_log("  Jefes Map usa claves TUPLA (1, 1). Adaptando...")
        for (eq, sec), jefe in jefes_map.items():
            jefes_data.append({'equipo': eq, 'sector': sec, 'jefe_lookup': jefe})
    
    df_jefes = pd.DataFrame(jefes_data)
    
    # Generate 'nombre_sector' key (E1S1 format) for Main Data
    try:
        # Main Data
        df_main['equipo'] = df_main['equipo'].astype(int)
        df_main['sector'] = df_main['sector'].astype(int)
        df_main['key_sector'] = "E" + df_main['equipo'].astype(str) + "S" + df_main['sector'].astype(str)
        
        # Jefes Data - only if we had tuple keys and need to generate key_sector
        if not is_string_key:
             df_jefes['equipo'] = df_jefes['equipo'].astype(int)
             df_jefes['sector'] = df_jefes['sector'].astype(int)
             df_jefes['key_sector'] = "E" + df_jefes['equipo'].astype(str) + "S" + df_jefes['sector'].astype(str)
        
    except Exception as e:
        print(f"Error generando llaves de sector: {e}")
        df_main['jefe_campo'] = "ERROR LLAVE"
        return df_main

    # Debug keys
    debug_log("DEBUG: --- MERGE KEYS CHECK ---")
    try:
        debug_log(f"Main DF Keys (Head): {df_main['key_sector'].head(5).tolist()}")
        debug_log(f"Main DF Keys (Tail): {df_main['key_sector'].tail(5).tolist()}")
        debug_log(f"Jefes DF Keys (Head): {df_jefes['key_sector'].head(5).tolist()}")
    except:
        pass
    
    # Check for intersection
    main_keys = set(df_main['key_sector'].unique())
    jefes_keys = set(df_jefes['key_sector'].unique())
    intersection = main_keys.intersection(jefes_keys)
    debug_log(f"DEBUG: Unique Main Keys: {len(main_keys)}")
    debug_log(f"DEBUG: Unique Jefes Keys: {len(jefes_keys)}")
    debug_log(f"DEBUG: Keys in Common: {len(intersection)}")
    if len(intersection) == 0:
        debug_log("DEBUG: CRITICAL - NO KEYS MATCH!")
    else:
        debug_log(f"DEBUG: Sample Match: {list(intersection)[0]}")

    # Merge (Left Join) on the generated string key
    df_merged = pd.merge(df_main, df_jefes[['key_sector', 'jefe_lookup']], on='key_sector', how='left')
    
    # Fill NaN with generic
    df_merged['jefe_lookup'] = df_merged['jefe_lookup'].fillna("POR ASIGNAR")
    
    # Rename/Assign to final column expected
    df_merged['jefe_campo'] = df_merged['jefe_lookup']
    
    return df_merged


def generate_excel(all_data, maestro, output_path):
    """Generate the output Excel file."""
    
    # Pre-process with Pandas for robust VLOOKUP
    if not all_data:
        return 0
        
    df = pd.DataFrame(all_data)
    # Note: 'jefe' might be in all_data if it came from previous logic, but we want to ensure it comes from our robust lookup
    # Actually, process_extraction passes 'jefes_campo' map to us, but signature here is (all_data, maestro, output_path).
    # Wait, previous signature was (all_data, maestro, jefes_campo, output_path).
    # I need to keep the signature consistent or update the caller. 
    # The user instruction asked for a modular function.
    # I will stick to the plan: process_extraction calls vincular, then passes enriched data here.
    # checking signature... oops, I need to see process_extraction to see how it calls this.
    
    wb_out = openpyxl.Workbook()
    ws = wb_out.active
    ws.title = "Riego Diario"

    # Styles
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )
    number_fmt = "#,##0"
    decimal_fmt = "#,##0.00"

    # New Order: Jefe de Campo (G) after Sector (F)
    headers = [
        "Fecha", "Dia", "Clave", "Nombre Sector", "Equipo", "Sector",
        "Jefe de Campo", # <--- NEW POSITION (Col 7 / G)
        "Variedad", "Tipo de Riego", "Has",
        "Horas", "m3/ha/hr", "M3",
        "M3/ha", "mm",
        "Fundo", "Caseta",
        # Old Jefe was here
        "Con Fertilizante",
        "Horario de Inicio"
    ]

    # Color definitions (Web-style palette)
    COLORS = {
        "Header": "364F6B",     # Dark Slate Blue (Professional Header)
        "Text": "FFFFFF",       # White Text for Header
        
        # Semantic Data Colors
        "Time": "FFF2CC",       # Soft Yellow/Orange for Time (Horas)
        "Land": "E2EFDA",       # Soft Green for Land (Has)
        "Water": "D9E1F2",      # Soft Blue for Water metrics (M3, mm, etc)
        "Input": "FAFAFA",      # Very light grey for input fields
        
        # Crop Specific (Pastels)
        "Olivo": "C6E0B4",      # Light Green
        "Olivo S. tree": "A9D08E", # Darker Green
        "Cerezo": "FFC7CE",     # Light Red/Pink
        "Avellano": "FFE699",   # Light Gold
        "Otro": "F2F2F2",       # Light Grey
        
        # Alerts
        "Fert": "FFEB9C",       # Yellow/Gold for Fertilizer Alert
        "FertText": "9C5700"    # Dark text for contrast
    }

    header_fill = PatternFill(start_color=COLORS["Header"], end_color=COLORS["Header"], fill_type="solid")
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True, size=11, color=COLORS["Text"])
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border

    # Sort by date, then equipo, then sector
    # all_data is a list of dicts. 
    all_data.sort(key=lambda x: (x['fecha'], x['equipo'], x['sector']))

    ws.sheet_view.showGridLines = False

    # Pre-define fills for performance
    fill_time = PatternFill(start_color=COLORS["Time"], end_color=COLORS["Time"], fill_type="solid")
    fill_land = PatternFill(start_color=COLORS["Land"], end_color=COLORS["Land"], fill_type="solid")
    fill_water = PatternFill(start_color=COLORS["Water"], end_color=COLORS["Water"], fill_type="solid")
    fill_fert = PatternFill(start_color=COLORS["Fert"], end_color=COLORS["Fert"], fill_type="solid")
    font_fert = Font(color=COLORS["FertText"], bold=True)

    row_num = 2
    for entry in all_data:
        eq = entry['equipo']
        sec = entry['sector']
        fecha = entry['fecha']
        hours = entry['horas']

        m = maestro.get((eq, sec), {})
        if not m:
            # Fallback: still output but with limited info
            m = {"clave": f"{eq}.{sec}", "has": 0, "variedad": "", "prod_est": "", "m3_ha_hr": 0}

        has = m["has"]
        m3_ha_hr = m["m3_ha_hr"]
        variedad = m["variedad"]

        # Use M3 from planilla if available, otherwise calculate
        if entry['m3_planilla'] and entry['m3_planilla'] > 0:
            m3 = entry['m3_planilla']
        else:
            m3 = hours * m3_ha_hr * has if has > 0 else 0

        m3_ha = m3 / has if has > 0 else 0
        mm = m3_ha / 10 if has > 0 else 0

        tipo_riego = get_tipo_riego(m) if m.get("variedad") else ""
        fundo, caseta = get_fundo_caseta(eq)
        dia_nombre = DIAS_SEMANA.get(datetime.combine(fecha, dtime()).weekday(), "")
        
        # Jefe comes from the enriched entry now
        jefe = entry.get('jefe_campo', 'POR ASIGNAR')
        nombre_sector = f"E{eq}S{sec}"

        # Crop color logic
        crop_color_hex = COLORS.get(tipo_riego, "FFFFFF")
        if "Olivo" in tipo_riego: crop_color_hex = COLORS["Olivo"]
        fill_crop = PatternFill(start_color=crop_color_hex, end_color=crop_color_hex, fill_type="solid")

        row_data = [
            fecha.strftime("%Y-%m-%d"), # A (1)
            dia_nombre,                 # B (2)
            m["clave"],                 # C (3)
            nombre_sector,              # D (4)
            eq,                         # E (5)
            sec,                        # F (6)
            jefe,                       # G (7) <--- NEW
            variedad,                   # H (8)
            tipo_riego,                 # I (9)
            has,                        # J (10)
            hours,                      # K (11)
            round(m3_ha_hr, 2),         # L (12)
            round(m3, 0),               # M (13)
            round(m3_ha, 2),            # N (14)
            round(mm, 2),               # O (15)
            fundo,                      # P (16)
            caseta,                     # Q (17)
            entry['con_fert'],          # R (18)
            "" # Horario de Inicio      # S (19)
        ]

        for col, val in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col, value=val)
            cell.border = thin_border
            
            # --- Coloring Logic ---
            # Adjusted indices because of new column insert at 7
            
            # Crop Info: Clave(3), Variedad(8), Tipo(9) -> Previously 3, 7, 8
            if col in (3, 8, 9): 
                cell.fill = fill_crop
            
            # Land (Has - Col 10) -> Prev 9
            elif col == 10:
                cell.fill = fill_land
                
            # Time (Horas - Col 11) -> Prev 10
            elif col == 11:
                cell.fill = fill_time

            # Water Metrics (Cols 12, 13, 14, 15) -> Prev 11, 12, 13, 14
            elif col in (12, 13, 14, 15):
                cell.fill = fill_water
            
            # Fertilizer Alert (Col 18) -> Prev 18 (Wait: 18? let's recount)
            # 16:Fundo, 17:Caseta, 18:Fert. Yes.
            elif col == 18 and val == "Si":
                cell.fill = fill_fert
                cell.font = font_fert
            
            # Jefe de Campo Alert (Col 7)
            elif col == 7 and val in ["POR ASIGNAR", "REVISAR MAESTRO"]:
                 cell.font = Font(color="FF0000", bold=True)

            # Alignment
            # Centers: Date(1), Dia(2), Clave(3), NomSec(4), Eq(5), Sec(6), Jefe(7), Fundo(16), Caseta(17), Fert(18)
            if col in (1, 2, 3, 4, 5, 6, 7, 16, 17, 18): 
                 cell.alignment = Alignment(horizontal="center", vertical="center")
            elif col in (10, 11, 12, 13, 14, 15): # Numbers
                 cell.alignment = Alignment(horizontal="right", vertical="center")
            
            # Number Formats
            # 13 is M3 (Integer)
            if col in (10, 11, 12, 14, 15): # Decimals
                cell.number_format = decimal_fmt
            elif col == 13: # M3 Integer
                cell.number_format = number_fmt

        row_num += 1

    # Column widths & Filter
    # A=12, B=5, C=8, D=10, E=7, F=7, G(Jefe)=20, H=18, I=16, J=8, K=8, L=9, M=10, N=9, O=8, P=7, Q=8, R=14, S=15
    col_widths = [12, 5, 8, 10, 7, 7, 20, 18, 16, 8, 8, 9, 10, 9, 8, 7, 8, 14, 15]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Hide columns
    # Hid: Clave(C), Variedad(H), Tipo(I), Has(J), m3/ha/hr(L), M3/ha(N), mm(O), Fundo(P), Caseta(Q)
    cols_to_hide = ['C', 'H', 'I', 'J', 'L', 'N', 'O', 'P', 'Q']
    for col_letter in cols_to_hide:
        ws.column_dimensions[col_letter].hidden = True

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{row_num - 1}"

    wb_out.save(output_path)
    return row_num - 2


def process_extraction(planilla_path, date_str, output_path, log_callback=print, manual_jefes_path=None):
    """
    Main extraction process.
    """
    log_callback(f"Procesando: {planilla_path}")
    log_callback(f"Rango Fechas: {date_str}")
    
    # 1. Load Jefes Mapping
    log_callback("Cargando maestro de Jefes de Campo...")
    # Open workbook first to check for internal 'Datos'
    try:
        wb = openpyxl.load_workbook(planilla_path, read_only=True, data_only=True)
    except Exception as e:
        log_callback(f"Error abriendo planilla entrada: {e}")
        return
        
    jefes_campo = load_jefes_campo(wb, base_path=planilla_path, manual_path=manual_jefes_path)
    
    if not jefes_campo:
        log_callback("ADVERTENCIA: No se pudo cargar ningun Jefe de Campo.")
        log_callback("  - Verifique que exista la hoja 'Datos' en el archivo de entrada.")
        log_callback("  - O que exista 'Planilla tipo Programacion riego.xlsx' en la carpeta.")
    else:
        log_callback(f"Maestro de Jefes cargado: {len(jefes_campo)} entradas.")

    target_dates = parse_dates(date_str)
    
    # 2. Extract Data
    log_callback("Extrayendo horas de riego de hojas EqX...")
    all_data = []
    log_callback(f"DEBUG: EQ_SHEETS tiene {len(EQ_SHEETS)} elementos: {EQ_SHEETS}")
    for sheet_name in EQ_SHEETS:
        entries = extract_from_eq_sheet(wb, sheet_name, target_dates)
        if entries:
            log_callback(f"  {sheet_name}: {len(entries)} registros")
        all_data.extend(entries)

    log_callback(f"  Total: {len(all_data)} registros con riego")

    wb.close()

    if len(all_data) == 0:
        log_callback("\nNo se encontraron datos de riego para las fechas indicadas.")
        return 0
    
    # --- VLOOKUP Logic ---
    log_callback(f"Asignando Jefes de Campo (VLOOKUP)...")
    try:
        df_jefes = pd.DataFrame(all_data)
        df_linked = vincular_jefes_campo(df_jefes, jefes_campo)
        # Convert back to list of dicts for generate_excel
        all_data_linked = df_linked.to_dict('records')
    except Exception as e:
        log_callback(f"Error en VLOOKUP Jefes: {e}. Usando datos sin procesar.")
        all_data_linked = all_data

    log_callback("Generando Excel de salida...")
    # Updated signature: generate_excel(all_data, maestro, output_path) - removed jefes_campo
    n_rows = generate_excel(all_data_linked, maestro={}, output_path=output_path)

    # 3. Sincronización Automática con Supabase
    try:
        df_final = pd.DataFrame(all_data_linked)
        # Necesitamos Fundo y Nombre Sector que se generan en generate_excel logic, 
        # pero aquí los reconstruimos rápido para el upload.
        def get_fundo_only(eq):
            fundo_map = {
                1: "DA", 2: "DA", 3: "DA", 4: "DA", 5: "DA", 6: "DA", 7: "DA", 9: "DA",
                10: "DJ", 11: "DJ", 12: "DJ", 13: "DJ", 14: "DJ", 15: "DJ", 16: "DJ", 17: "DJ",
                18: "DJ", 19: "DJ", 20: "DJ", 21: "DJ", 22: "DJ"
            }
            return fundo_map.get(eq, "")

        df_final['Fundo'] = df_final['equipo'].apply(get_fundo_only)
        df_final['Nombre Sector'] = "E" + df_final['equipo'].astype(str) + "S" + df_final['sector'].astype(str)
        df_final['Fecha'] = df_final['fecha']
        df_final['Horas'] = df_final['horas']
        df_final['M3'] = df_final['m3_planilla'] # O el calculado si m3_planilla es None, simplificamos:
        df_final['Con Fertilizante'] = df_final['con_fert']

        upload_to_supabase(df_final, log_callback)
    except Exception as e:
        import traceback
        error_msg = traceback.format_exc()
        log_callback(f"Error detallado Supabase: {error_msg}")
        log_callback(f"Error preparando datos para Supabase: {e}")

    log_callback(f"\nListo! {n_rows} filas generadas en: {output_path}")
    return n_rows


def main():
    if len(sys.argv) < 4:
        print(__doc__)
        sys.exit(1)

    planilla_path = sys.argv[1]
    date_str = sys.argv[2]
    output_path = sys.argv[3]

    try:
        process_extraction(planilla_path, date_str, output_path)
    except Exception as e:
        print(f"Error: {e}")
        sys.exit(1)


if __name__ == "__main__":
    main()
