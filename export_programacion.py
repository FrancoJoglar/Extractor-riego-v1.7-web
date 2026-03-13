import pandas as pd
from supabase import create_client, Client
import os
from datetime import datetime, timedelta
import tkinter as tk
from tkinter import filedialog

# Configuración de Supabase
SUPABASE_URL = "https://miikjrfqmmkzknyngwen.supabase.co"
SUPABASE_KEY = "sb_publishable_tixCrY6poQrVDEzeRuDhow_ND7tvVCQ" 
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)

def sincronizar_a_supabase(fecha_filtro=None):
    """Sincroniza los datos procesados con Supabase."""
    from tkinter import filedialog
    
    print("Iniciando sincronización con Supabase...")
    
    root = tk.Tk()
    root.withdraw()
    root.attributes('-topmost', True)
    
    input_path = filedialog.askopenfilename(
        title="Seleccionar Planilla Procesada",
        filetypes=[("Excel Files", "*.xlsx *.xls")]
    )
    
    root.destroy()
    root.quit()
    
    if not input_path:
        print("Sincronización cancelada.")
        return
    
    try:
        df = pd.read_excel(input_path)
        print(f"Planilla cargada: {input_path}")
    except Exception as e:
        print(f"Error al leer la planilla: {e}")
        return
    
    from extract_riego import upload_to_supabase
    
    try:
        upload_to_supabase(df, print)
    except Exception as e:
        print(f"Error al sincronizar: {e}")


def exportar_para_programacion(fecha_filtro=None, output_path=None):
    print("Obteniendo datos de 'vista_riegos_solicitados'...")
    
    # Obtener datos pendientes (asumimos que solo queremos programar los pendientes)
    try:
        query = supabase.table("vista_riegos_solicitados").select("*").eq("estado", "pendiente")
        
        # Aplicar filtro de fecha si existe (formato DD-MM-YYYY)
        if fecha_filtro:
            print(f"Filtrando por fecha: {fecha_filtro}")
            query = query.eq("fecha_solicitado", fecha_filtro)
            
        response = query.execute()
        data = response.data
    except Exception as e:
        print(f"Error al conectar con Supabase: {e}")
        return

    if not data:
        print(f"No hay riegos pendientes para programar{' para la fecha ' + fecha_filtro if fecha_filtro else ''}.")
        return

    df = pd.DataFrame(data)
    
    # Convertir fecha_solicitado a datetime para ordenar correctamente si es string
    if 'fecha_solicitado' in df.columns:
        df['fecha_solicitado'] = pd.to_datetime(df['fecha_solicitado'], dayfirst=True, errors='coerce')

    # Extraer número de equipo para ordenar correctamente (1, 2, 10, etc.)
    if 'equipo_nombre' in df.columns:
        # Extraer solo los digitos y convertir a entero
        df['equipo_num'] = df['equipo_nombre'].str.extract(r'(\d+)').astype(float)

    # Ordenar por Fecha, Numero de Equipo y Sector
    sort_cols = []
    if 'fecha_solicitado' in df.columns: sort_cols.append('fecha_solicitado')
    
    # Usar el número de equipo si existe, sino el nombre
    if 'equipo_num' in df.columns: 
        sort_cols.append('equipo_num')
    elif 'equipo_nombre' in df.columns: 
        sort_cols.append('equipo_nombre')
        
    if 'sector_nombre' in df.columns: sort_cols.append('sector_nombre')
    
    if sort_cols:
        df = df.sort_values(by=sort_cols)
        
    # Eliminar columna temporal de ordenamiento
    if 'equipo_num' in df.columns:
        df = df.drop(columns=['equipo_num'])

    # Volver a formato string para el excel
    if 'fecha_solicitado' in df.columns:
        df['fecha_solicitado'] = df['fecha_solicitado'].dt.strftime('%d-%m-%Y')

    # Redondear m3_estimados y convertir a entero
    if 'm3_estimados' in df.columns:
        df['m3_estimados'] = df['m3_estimados'].fillna(0).round(0).astype(int)

    # Agregar columna para 'Hora de Inicio' vacía (o con un valor por defecto si quisiéramos)
    # Aquí es donde iría la lógica automática más adelante
    df['Hora de Inicio'] = "" 
    
    # Selección y orden de columnas para el reporte
    columnas_deseadas = [
        'fecha_solicitado', 
        'equipo_nombre', 
        'sector_nombre', 
        'jefe_campo', 
        'horas_solicitadas', 
        'm3_estimados', 
        'con_fertilizante',
        'Hora de Inicio' # Columna nueva manual al final
    ]
    
    # Filtrar solo las que existen en el dataframe
    cols_final = [c for c in columnas_deseadas if c in df.columns or c in ['Hora de Inicio']]
    df = df[cols_final]

    # Renombrar columnas para el reporte final (Mayúscula inicial)
    nombre_columnas = {
        'fecha_solicitado': 'Fecha Solicitado',
        'equipo_nombre': 'Equipo',
        'sector_nombre': 'Sector',
        'jefe_campo': 'Jefe de Campo',
        'horas_solicitadas': 'Horas',
        'm3_estimados': 'M3 Estimados',
        'con_fertilizante': 'Con Fertilizante',
        'Hora de Inicio': 'Hora Inicio'
    }
    df = df.rename(columns=nombre_columnas)

    # Generar nombre de archivo
    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_filename = f"Planilla_Programacion_{timestamp}.xlsx"
        output_path = os.path.join(os.getcwd(), output_filename)
    else:
        output_dir = os.path.dirname(output_path)
        if output_dir and not os.path.exists(output_dir):
            os.makedirs(output_dir, exist_ok=True)
    
    try:
        # Guardar archivo base
        df.to_excel(output_path, index=False)
        
        # Aplicar estilos con openpyxl
        try:
            import openpyxl
            from openpyxl.styles import PatternFill, Font, Alignment
            
            wb = openpyxl.load_workbook(output_path)
            ws = wb.active
            
            # Estilos: Fondo Azul, Letra Blanca Negrita
            azul_encabezado = "0070C0"
            header_fill = PatternFill(start_color=azul_encabezado, end_color=azul_encabezado, fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            # Color naranjo para Hora Inicio
            naranjo = "FF6600"
            hora_inicio_font = Font(color=naranjo, bold=True)
            
            # Aplicar a la primera fila (encabezados)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal='center', vertical='center')

            # Identificar grupos de equipos con colores de fondo sutiles
            # Paleta de colores en tonos azules extremadamente claros para alternar
            colores_fondo = ["F2F7FB", "E1EEF8"] # Azules ultra claros sutiles            
            equipo_col_idx = None
            # Buscar la columna 'Equipo'
            for idx, cell in enumerate(ws[1], 1):
                if cell.value == "Equipo":
                    equipo_col_idx = idx
                    break
            
            if equipo_col_idx:
                current_color_idx = 0
                last_equipo = None
                
                for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row)):
                    val_equipo = row[equipo_col_idx-1].value
                    
                    # Solo cambiar color si cambia el equipo y no es la primera fila
                    if i > 0 and val_equipo != last_equipo:
                         current_color_idx = (current_color_idx + 1) % len(colores_fondo)
                    
                    last_equipo = val_equipo
                    
                    row_fill = PatternFill(start_color=colores_fondo[current_color_idx], end_color=colores_fondo[current_color_idx], fill_type="solid")
                    
                    for idx, cell in enumerate(row, 1):
                        cell.fill = row_fill
                        cell.alignment = Alignment(horizontal='center', vertical='center')
                        
                        # Resaltar Sector (3) y M3 Estimados (6) con negrita
                        if idx in [3, 6]:
                            cell.font = Font(bold=True)
                        
                        # Hora Inicio (8) en color naranjo
                        if idx == 8:
                            cell.font = hora_inicio_font

            # Ajustar ancho de columnas automáticamente (aproximado)
            for column in ws.columns:
                max_length = 0
                column = [cell for cell in column]
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = (max_length + 2)
                ws.column_dimensions[column[0].column_letter].width = adjusted_width

            wb.save(output_path)
            print("Estilos aplicados correctamente.")
            
        except ImportError:
            print("AVISO: 'openpyxl' no está instalado. El archivo se generó sin estilos.")
        except Exception as e:
            print(f"Error al aplicar estilos: {e}")

        print(f"Planilla generada exitosamente: {output_path}")
        os.startfile(output_path) # Intenta abrir el archivo automáticamente (solo Windows)
    except Exception as e:
        print(f"Error al guardar el archivo Excel: {e}")

if __name__ == "__main__":
    # Filtrar automáticamente para el día de mañana
    manana = (datetime.now() + timedelta(days=1)).strftime('%d-%m-%Y')
    print(f"Ejecutando exportación para la fecha: {manana}")
    exportar_para_programacion(manana)
