"""
================================================================================
SCRIPT: programar_horarios.py (NUEVA VERSIÓN)
================================================================================
PROPÓSITO: 
    Calcula horarios de riego para equipos y sectores con la nueva lógica:
    
    1. Sectores CON fertilización: SIEMPRE desde 06:00, hasta max 16:00
    2. Sectores SIN fertilización: cualquier horario, después de los CON fert
    3. Equipos >14 horas: cálculo hacia atrás desde 06:00 para SIN fert
    4. NO solapar: cada sector inicia donde termina el anterior

================================================================================
"""

import pandas as pd
from supabase import create_client, Client
import os
from datetime import datetime

# ================================================================================
# CONFIGURACIÓN - CREDENCIALES SUPABASE
# ================================================================================
SUPABASE_URL = "https://miikjrfqmmkzknyngwen.supabase.co"
SUPABASE_KEY = "sb_publishable_tixCrY6poQrVDEzeRuDhow_ND7tvVCQ"
supabase: Client = create_client(SUPABASE_URL, SUPABASE_KEY)


# ================================================================================
# FUNCIONES AUXILIARES
# ================================================================================

def formatear_hora(hora_float):
    """Convierte hora float (6.5) a string HH:MM"""
    horas = int(hora_float)
    minutos = int(round((hora_float - horas) * 60))
    # Ajustar si pasa de 24 horas
    if horas >= 24:
        horas = horas - 24
    # 00:00 funciona como desactivador, usar 00:01 mínimo
    if horas == 0 and minutos == 0:
        minutos = 1
    return f"{horas:02d}:{minutos:02d}"


def es_fertilizante(val):
    """Convierte valor a booleano (maneja strings, bools, etc.)"""
    if isinstance(val, bool):
        return val
    if isinstance(val, (int, float)):
        return bool(val)
    if isinstance(val, str):
        return val.upper() in ['TRUE', 'VERDADERO', '1', 'YES', 'S', 'SI']
    return bool(val)


def extraer_numeros(df):
    """Extrae números de equipo y sector"""
    if 'equipo_nombre' in df.columns:
        df['equipo_num'] = df['equipo_nombre'].str.extract(r'(\d+)').astype(float)
    if 'sector_nombre' in df.columns:
        df['sector_num'] = df['sector_nombre'].str.extract(r'S(\d+)').astype(float)
    return df


# ================================================================================
# FUNCIÓN PRINCIPAL: calcular_horario(equipo_df)
# ================================================================================
def calcular_horario(equipo_df):
    """
    Calcula horarios para un equipo con la nueva lógica:
    
    REGLAS:
    1. Sectores CON fert: 06:00 - 16:00 (SECUENCIAL)
    2. Sectores SIN fert: cualquier horario después de los CON fert
    3. Si total > 14h: cálculo hacia atrás desde 06:00 para SIN fert
    4. NO solapar
    """
    resultados = {}
    hora_termino_fert = 6.0  # Valor por defecto
    
    # Convertir fertilizante a booleano
    equipo_df = equipo_df.copy()
    equipo_df['con_fertilizante'] = equipo_df['con_fertilizante'].apply(es_fertilizante)
    
    # Calcular total de horas del equipo
    total_horas = equipo_df['horas_solicitadas'].sum()
    
    # Separar sectores: CON fert vs SIN fert
    sectores_con_fert = equipo_df[equipo_df['con_fertilizante'] == True].sort_values('sector_num')
    sectores_sin_fert = equipo_df[equipo_df['con_fertilizante'] == False].sort_values('sector_num')
    
    tiene_fertilizante = len(sectores_con_fert) > 0
    
    print(f"  [INFO] Total horas: {total_horas}h | CON fert: {len(sectores_con_fert)} | SIN fert: {len(sectores_sin_fert)}")
    
    # ============================================================================
    # 1. PROGRAMAR CON FERTILIZACIÓN (PRIMERO - SIEMPRE DESDE 06:00)
    # ============================================================================
    if len(sectores_con_fert) > 0:
        # Asignar horas a cada sector
        horas_por_sector = []
        for idx, row in sectores_con_fert.iterrows():
            horas_por_sector.append({'idx': idx, 'horas': row['horas_solicitadas'], 'sector': row['sector_nombre']})
        
        # Calcular tiempo total de CON fert
        total_horas_fert = sum(h['horas'] for h in horas_por_sector)
        
        # Programar secuencialmente desde 06:00
        hora_actual = 6.0
        for item in horas_por_sector:
            hora_fin = hora_actual + item['horas']
            tipo = "Horario Inicio" if hora_actual == 6.0 else "Secuencial"
            
            resultados[item['idx']] = {
                'Hora Inicio': formatear_hora(hora_actual),
                'Tipo': tipo
            }
            print(f"    [CON FERT] {item['sector']}: {formatear_hora(hora_actual)} - {item['horas']}h (fin: {formatear_hora(hora_fin)})")
            hora_actual += item['horas']
        
        # Guardar dónde terminó el último CON fert
        hora_termino_fert = hora_actual
    
    # ============================================================================
    # 2. PROGRAMAR SIN FERTILIZACIÓN (DESPUÉS)
    # ============================================================================
    if len(sectores_sin_fert) > 0:
        # Determinar hora de inicio para SIN fert
        if tiene_fertilizante:
            # Continuar desde donde terminó CON fert
            hora_actual = hora_termino_fert
        elif total_horas > 14:
            # Equipos >14h: cálculo hacia atrás desde 06:00
            # El primer sector SIN fert empieza en 06:00 - primera_hora
            primera_hora = sectores_sin_fert['horas_solicitadas'].iloc[0]
            hora_actual = 6.0 - primera_hora
            if hora_actual < 0:
                hora_actual = 0  # mínimo media noche
        else:
            # Equipos <=14h: iniciar a las 06:00
            hora_actual = 6.0
        
        primer_sector = len(resultados) == 0
        
        for idx, row in sectores_sin_fert.iterrows():
            horas_req = row['horas_solicitadas']
            hora_termino = hora_actual + horas_req
            
            # Si pasa de 24h, ajustar (nuevo día)
            if hora_termino > 24:
                hora_actual = hora_actual  # continúa al día siguiente
            
            resultados[idx] = {
                'Hora Inicio': formatear_hora(hora_actual),
                'Tipo': 'Horario Inicio' if primer_sector else 'Secuencial'
            }
            
            print(f"    [SIN FERT] {row['sector_nombre']}: {formatear_hora(hora_actual)} - {horas_req}h")
            
            # Avanzar secuencialmente
            hora_actual += horas_req
            primer_sector = False
    
    return resultados


# ================================================================================
# FUNCIÓN: obtener_datos(fecha_filtro=None)
# ================================================================================
def obtener_datos(fecha_filtro=None):
    """Obtiene riegos pendientes de Supabase"""
    print("Obteniendo datos de riegos pendientes...")
    
    query = supabase.table("vista_riegos_solicitados").select("*").eq("estado", "pendiente")
    
    if fecha_filtro:
        print(f"Filtrando por fecha: {fecha_filtro}")
        query = query.eq("fecha_solicitado", fecha_filtro)
    
    response = query.execute()
    data = response.data
    
    if not data:
        print(f"No hay riegos para la fecha {fecha_filtro}")
        return None
    
    df = pd.DataFrame(data)
    print(f"Se encontraron {len(df)} riegos")
    return df


# ================================================================================
# FUNCIÓN: programar_por_fecha_equipo(df)
# ================================================================================
def programar_por_fecha_equipo(df):
    """Itera por fechas y equipos calculando horarios"""
    df = df.copy()
    df['Hora Inicio'] = ""
    df['Tipo Programación'] = ""
    df['hora_inicio_num'] = 0.0  # Columna auxiliar para ordenar
    
    # Extraer números
    df = extraer_numeros(df)
    
    if 'fecha_solicitado' not in df.columns or 'equipo_num' not in df.columns:
        print("ERROR: Faltan columnas necesarias")
        return df
    
    fechas = df['fecha_solicitado'].unique()
    
    for fecha in sorted(fechas):
        fecha_df = df[df['fecha_solicitado'] == fecha]
        equipos = fecha_df['equipo_num'].unique()
        
        for equipo in sorted(equipos):
            equipo_df = fecha_df[fecha_df['equipo_num'] == equipo]
            
            # Nombre del equipo
            nombre_equipo = equipo_df['equipo_nombre'].iloc[0]
            print(f"\n[EQUIPO {int(equipo)}] {nombre_equipo}")
            
            # Calcular horarios
            resultados = calcular_horario(equipo_df)
            
            # Asignar resultados al DataFrame
            for idx, row in equipo_df.iterrows():
                if idx in resultados:
                    hora_str = resultados[idx]['Hora Inicio']
                    df.loc[idx, 'Hora Inicio'] = hora_str
                    df.loc[idx, 'Tipo Programación'] = resultados[idx]['Tipo']
                    # Convertir hora a número para ordenar (HH:MM -> float)
                    try:
                        h, m = hora_str.split(':')
                        df.loc[idx, 'hora_inicio_num'] = int(h) + int(m)/60
                    except:
                        pass
    
    # ORDENAR por hora de inicio DENTRO de cada equipo
    df = df.sort_values(by=['fecha_solicitado', 'equipo_num', 'hora_inicio_num'])
    
    # Eliminar columna auxiliar
    df = df.drop(columns=['hora_inicio_num'], errors='ignore')
    
    return df


# ================================================================================
# FUNCIÓN: guardar_excel(df, output_path)
# ================================================================================
def guardar_excel(df, output_path):
    """Guarda DataFrame en Excel con estilos"""
    # Eliminar columnas auxiliares
    cols_a_borrar = ['equipo_num', 'sector_num']
    df = df.drop(columns=[c for c in cols_a_borrar if c in df.columns], errors='ignore')
    
    # Formatear fecha
    if 'fecha_solicitado' in df.columns:
        if isinstance(df['fecha_solicitado'].iloc[0], (datetime, pd.Timestamp)):
            df['fecha_solicitado'] = df['fecha_solicitado'].dt.strftime('%d-%m-%Y')
    
    # Redondear M3
    if 'm3_estimados' in df.columns:
        df['m3_estimados'] = df['m3_estimados'].fillna(0).round(0).astype(int)
    
    # Renombrar columnas
    nombres = {
        'fecha_solicitado': 'Fecha Solicitado',
        'equipo_nombre': 'Equipo',
        'sector_nombre': 'Sector',
        'jefe_campo': 'Jefe de Campo',
        'horas_solicitadas': 'Horas',
        'm3_estimados': 'M3 Estimados',
        'con_fertilizante': 'Con Fertilizante',
        'Hora Inicio': 'Hora Inicio',
        'Tipo Programación': 'Tipo Programación'
    }
    df = df.rename(columns=nombres)
    
    # Ordenar columnas
    columnas_orden = [c for c in ['Fecha Solicitado', 'Equipo', 'Sector', 'Jefe de Campo', 'Horas', 'M3 Estimados', 'Con Fertilizante', 'Hora Inicio', 'Tipo Programación'] if c in df.columns]
    df = df[columnas_orden]
    
    # Guardar
    df.to_excel(output_path, index=False)
    
    # ============================================================
    # APLICAR ESTILOS A LA PLANILLA DE EXPORTACIÓN
    # ============================================================
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment
        from openpyxl.utils import get_column_letter
        
        wb = openpyxl.load_workbook(output_path)
        ws = wb.active
        
        # ============================================
        # 1. CONFIGURACIÓN DE COLORES Y ESTILOS
        # ============================================
        
        # Encabezado: Azul oscuro #1F4E78, texto blanco, negrita
        AZUL_ENCABEZADO = "1F4E78"
        header_fill = PatternFill(start_color=AZUL_ENCABEZADO, end_color=AZUL_ENCABEZADO, fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        
        # Paleta de colores pastel por equipo (suave y legible)
        # Cada equipo tendrá un color diferente de esta paleta
        PALETA_COLORES_EQUIPOS = [
            "B4D7FF",  # Celeste pastel
            "D4E5F7",  # Azul muy claro
            "E8F1F8",  # Gris azulado claro
            "CCE5FF",  # Azul cielo
            "E0EACD",  # Verde pastel
            "F5DEB3",  # Dorado pastel
            "FFDAC1",  # Durazno
            "E2D5CF",  # Lila pastel
            "D7E8D5",  # Verde minta
            "F0E68C",  # Amarillo khaki
            "B0E0E6",  # Turquesa pastel
            "FFB6C1",  # Rosa claro
            "E6E6FA",  # Lavanda
            "FFFACD",  # Limón
            "F5F5DC",  # Beige
            "C1E1C1",  # Verde té
            "CFE2F3",  # Azul hielo
            "FCE5CD",  # Naranja pastel
            "D9EAD3",  # Verde menta
            "E8DAEF",  # Púrpura pastel
            "F3E5AB",  # Crema
            "CEDBD9",  # Verde grisáceo
        ]
        
        # Crear mapa de colores por equipo
        equipos_unicos = df['Equipo'].unique() if 'Equipo' in df.columns else []
        color_por_equipo = {}
        for i, eq in enumerate(sorted(equipos_unicos)):
            color_por_equipo[eq] = PALETA_COLORES_EQUIPOS[i % len(PALETA_COLORES_EQUIPOS)]
        
        # Fuentes
        font_dato = Font(name="Calibri", size=11, bold=False)
        font_negrita = Font(name="Calibri", size=11, bold=True)
        font_importante = Font(name="Calibri", size=12, bold=True)  # Sector, M3
        font_hora = Font(name="Calibri", size=12, bold=True, color="FF6600")  # Hora Inicio en NARANJO
        
        # ============================================
        # 2. MAPEO DE COLUMNAS
        # ============================================
        # Mapeo de nombres de columnas a índices (0-based)
        columnas = {cell.value: idx for idx, cell in enumerate(ws[1])}
        
        # Columnas que deben estar en NEGRITA
        col_sector = columnas.get('Sector', None)
        col_m3 = columnas.get('M3 Estimados', columnas.get('Volumen', None))
        col_hora = columnas.get('Hora Inicio', columnas.get('Hora de inicio', None))
        
        columnas_negrita = {col_sector, col_m3, col_hora}
        
        # Columna Equipo para grouping (índice)
        col_equipo = columnas.get('Equipo', None)
        
        # ============================================
        # 3. APLICAR ESTILOS AL ENCABEZADO
        # ============================================
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # ============================================
        # 4. APLICAR ESTILOS A LAS FILAS DE DATOS
        # ============================================
        equipo_actual = None
        
        for row in ws.iter_rows(min_row=2):
            # Obtener valor del equipo para esta fila
            if col_equipo is not None:
                equipo = row[col_equipo].value
            else:
                equipo = None
            
            # Si cambia el equipo, obtener nuevo color
            if equipo != equipo_actual:
                equipo_actual = equipo
            
            # Obtener color para este equipo
            fill_color = color_por_equipo.get(equipo, "FFFFFF")
            fill = PatternFill(start_color=fill_color, end_color=fill_color, fill_type="solid")
            
            # Aplicar estilo a cada celda
            for col_idx, cell in enumerate(row):
                cell.fill = fill
                cell.alignment = Alignment(horizontal='center', vertical='center')
                
                # Hora Inicio en NARANJO negrita
                if col_idx == col_hora:
                    cell.font = font_hora
                # Columnas importantes en negrita (Sector, M3)
                elif col_idx in columnas_negrita and col_idx is not None:
                    cell.font = font_importante
                else:
                    cell.font = font_dato
        
        # ============================================
        # 5. AJUSTAR ANCHO DE COLUMNAS
        # ============================================
        anchos = {
            'A': 18,  # Fecha
            'B': 12,  # Equipo
            'C': 10,  # Sector
            'D': 15,  # Jefe de Campo
            'E': 8,   # Horas
            'F': 14,  # M3 Estimados
            'G': 18,  # Con Fertilizante
            'H': 14,  # Hora Inicio
            'I': 18,  # Tipo Programación
        }
        
        for col_letter, width in anchos.items():
            if col_letter in ws.column_dimensions:
                ws.column_dimensions[col_letter].width = width
        
        # Guardar
        wb.save(output_path)
        print("Estilos aplicados - FORMATO CORPORATIVO COMPLETO")
        
    except Exception as e:
        print(f"Error aplicando estilos: {e}")
        print(f"Error estilos: {e}")
    
    print(f"\nArchivo guardado: {output_path}")
    os.startfile(output_path)


# ================================================================================
# FUNCIÓN PRINCIPAL: programar_horarios(fecha, output_path)
# ================================================================================
def programar_horarios(fecha_filtro=None, output_path=None):
    """Función para ser llamada desde la GUI"""
    print(f"\n{'='*60}")
    print(f"INICIO PROGRAMACIÓN DE HORARIOS - Fecha: {fecha_filtro}")
    print(f"{'='*60}\n")
    
    # 1. Obtener datos
    df = obtener_datos(fecha_filtro)
    
    if df is None or df.empty:
        print("No hay datos para programar")
        return
    
    # 2. Convertir fechas
    if 'fecha_solicitado' in df.columns:
        df['fecha_solicitado'] = pd.to_datetime(df['fecha_solicitado'], dayfirst=True, errors='coerce')
    
    # 3. Extraer números de equipo y sector
    df = extraer_numeros(df)
    
    # 4. Ordenar
    df = df.sort_values(by=['fecha_solicitado', 'equipo_num', 'sector_num'])
    
    # 4. Calcular horarios
    df_programado = programar_por_fecha_equipo(df)
    
    # 5. Guardar
    if output_path is None:
        script_dir = os.path.dirname(os.path.abspath(__file__))
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = os.path.join(script_dir, f"Planilla_Programacion_{timestamp}.xlsx")
    
    guardar_excel(df_programado, output_path)
    
    print(f"\n{'='*60}")
    print("PROGRAMACIÓN COMPLETADA")
    print(f"{'='*60}\n")


# ================================================================================
# PUNTO DE ENTRADA
# ================================================================================
if __name__ == "__main__":
    import sys
    fecha = sys.argv[1] if len(sys.argv) > 1 else None
    programar_horarios(fecha)
